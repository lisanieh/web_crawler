import requests
from bs4 import BeautifulSoup
from selenium import webdriver
import time
import openpyxl
import re

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

#settings
scroll_time = int(input("捲動次數:"))
driver = webdriver.Chrome()
url = "https://udn.com/news/breaknews/1/99#breaknews"
driver.get(url)


#模擬網頁滾動事件
for i in range(1, scroll_time+1):
    time.sleep(3) #延遲執行時間
    print(f"now scroll {i}/{scroll_time}")
    js = "window.scrollTo(0, document.body.scrollHeight);"
    driver.execute_script(js)

#檢查是否滑到底了


#抓取html資料
time.sleep(3) #延遲執行時間
r = driver.page_source
soup = BeautifulSoup(r,"html.parser")
head = soup.select("div.story-list__text > h2 > a")
print("article number =", len(head))
print("-------------------")

#counting settings
count = 0
i = count + 1

#output in excel
workbook = openpyxl.Workbook()
worksheet = workbook.active

#抓取分頁資料，並存入excel中
for h in head: 
    count += 1
    print("on the process of ",count)
    #爬取標題
    print("appending head")
    worksheet.cell(row = i,column = 1,value = h.text) # j = 1 means header
    # print(count,h.text) #h.text為a的內文
    #進入內文+爬取url
    url = "https://udn.com" + h["href"] #url更新為內文的網址
    print("appending url")
    worksheet.cell(row = i,column = 2,value = url) # j = 2 means url
    # print("url : ",url)
    r = requests.get(url)
    soup = BeautifulSoup(r.text,"html.parser")

    #爬取記者
    find_reporter = soup.select("span.article-content__author a")
    tmp = ""
    for d in find_reporter: #可能會有多個記者
        tmp = tmp + d.text + " "
    print("appending reporter")
    worksheet.cell(row = i,column = 3,value = tmp) # j = 3 means reporter
    # print(count,"reporter : ",tmp)
    tmp = None #free tmp
    #爬取時間
    find_time = soup.select("time.article-content__time")
    tmp = ''
    for d in find_time:
        tmp = tmp + d.text + " "
    print("appending time")
    worksheet.cell(row = i,column = 4,value = tmp) # j = 4 means time
    # print("time : ",tmp)
    tmp = None #free tmp
    #爬取報社
    find_report = soup.select("span.article-content__author")
    for d in find_report:
        print("appending report location")
        worksheet.cell(row = i,column = 5,value = d.text) # j = 5 means location
        # print("報社 : ",i.text)
    #爬取內文
    find_article = soup.select("section.article-content__editor p")
    tmp = ''
    for d in find_article:
        tmp = tmp + d.text + "\n"
    print("appending article")
    text = ILLEGAL_CHARACTERS_RE.sub(r'',tmp)
    worksheet.cell(row = i,column = 6,value = text) # j = 6 means article
    # print("內文 : ",i.text)
    tmp = None 
    print("-------------------")
    workbook.save("output.xlsx")
    i += 1
    
print("repti finished")